import os
import pandas as pd
import io
import zipfile
import json
import openpyxl
import numpy as np
import streamlit as st
from sqlalchemy import create_engine, text, inspect

# --- BLOCCO DI INIZIALIZZAZIONE DELLO STATO ---
# Questo blocco rende la pagina autosufficiente
if 'wizard_step' not in st.session_state:
    st.session_state['wizard_step'] = 0
if 'tipo_struttura' not in st.session_state:
    st.session_state.tipo_struttura = 'Ditta'
# --- FINE BLOCCO ---

# --- CONFIGURAZIONE E HELPER ---
BASE_DIR = os.getcwd() 
DATA_DIR = os.path.join(BASE_DIR, 'data')
DB_PATH = os.path.join(BASE_DIR, 'db', 'imported_data.sqlite')
MAPPING_BASE_DIR = os.path.join(BASE_DIR, 'mapping')
EXPORT_BASE_DIR = os.path.join(BASE_DIR, 'export')

# VERSIONE FINALE DELLA FUNZIONE HELPER - Supporta la logica "ibrida"
def crea_righe_multiple(df: pd.DataFrame, key_cols_map: dict, context_cols_map: dict, unpivot_map: dict) -> pd.DataFrame:
    """
    Crea righe multiple con logica ibrida: la prima riga è completa, le successive sono sparse.
    - key_cols_map: Colonne chiave da ripetere su OGNI riga. {dest: source}
    - context_cols_map: Colonne di contesto da mostrare SOLO sulla prima riga. {dest: source}
    - unpivot_map: Colonne da trasformare. {dest: [source_1, source_2, ...]}
    """
    final_rows = []
    num_groups = max(len(v) for v in unpivot_map.values()) if unpivot_map else 0
    if num_groups == 0: return pd.DataFrame()

    for _, source_row in df.iterrows():
        is_first_row_for_this_company = True
        
        # Prepara i dati chiave che si ripeteranno sempre
        key_data = {dest_col: source_row.get(source_col) for dest_col, source_col in key_cols_map.items()}
        # Prepara i dati di contesto che appariranno solo una volta
        context_data = {dest_col: source_row.get(source_col) for dest_col, source_col in context_cols_map.items()}

        for i in range(num_groups):
            new_row_segment = {}
            is_valid_row = False
            
            for dest_col, source_cols_list in unpivot_map.items():
                try:
                    value = source_row.get(source_cols_list[i], '')
                    new_row_segment[dest_col] = value
                    if str(value).strip():
                        is_valid_row = True
                except IndexError:
                    new_row_segment[dest_col] = ''
            
            if is_valid_row:
                if is_first_row_for_this_company:
                    # Per la prima riga, unisci tutto: chiavi + contesto + dati trasformati
                    full_row = {**key_data, **context_data, **new_row_segment}
                    is_first_row_for_this_company = False
                else:
                    # Per le righe successive, unisci solo: chiavi + dati trasformati
                    full_row = {**key_data, **new_row_segment}
                
                final_rows.append(full_row)
                
    return pd.DataFrame(final_rows)

try:
    engine = create_engine(f'sqlite:///{DB_PATH}')
except Exception as e:
    st.error(f"Errore critico nel motore del database: {e}"); st.stop()

# Sostituisci questa funzione
def load_template_callback(config):
    """
    Callback per caricare i dati di un template selezionato nello stato della sessione.
    """
    mode = config['mode']
    template_name = st.session_state.get(f"template_loader_{mode}")
    
    # Chiave per il nostro stato di mappatura centrale
    # Aggiornata per la nuova struttura: live_mapping_state_key non è più usata direttamente così
    # Ora lo stato è per-tabella, quindi resetteremo solo il loaded_template_data
    
    # Reset dei nomi dei template
    st.session_state.loaded_template_name = None
    if 'loaded_template_data' in st.session_state:
        del st.session_state['loaded_template_data']

    # Pulisce lo stato dei widget e lo stato centrale
    # Non è più necessario pulire map_global_* qui, perché lo stato è per-tabella
    # e sarà gestito dal selettore della tabella di struttura.

    if not template_name or template_name == "-- Non caricare nulla --":
        return

    # Costruisce il percorso e carica il file
    templates_dir = os.path.join(config["mapping_dir"], "templates")
    safe_filename = "".join(c for c in template_name if c.isalnum() or c in (' ', '_')).rstrip()
    template_path = os.path.join(templates_dir, f"{safe_filename.replace(' ', '_')}.json")

    if os.path.exists(template_path):
        with open(template_path, 'r', encoding='utf-8') as f:
            # Carica TUTTI i dati del template in uno stato dedicato
            st.session_state.loaded_template_data = json.load(f)
            st.session_state.loaded_template_name = template_name 
            # Non facciamo altro qui. L'inizializzazione dello stato avverrà in step_5
    else:
        st.error(f"File del template '{template_name}' non trovato.")

def on_mode_change():
    """Pulisce lo stato della sessione che dipende dalla modalità quando questa viene cambiata."""
    keys_to_clear = [k for k in st.session_state.keys() if '_ms_' in k or 'tabella_in_modifica' in k or 'mass_edits' in k or 'exported_file_paths' in k or k.startswith('live_mapping_state_')]
    for key in keys_to_clear:
        if key in st.session_state: del st.session_state[key]

def delete_file(directory, filename):
    """Funzione helper per cancellare un file in modo sicuro."""
    try:
        filepath = os.path.join(directory, filename)
        if os.path.exists(filepath): os.remove(filepath)
    except Exception as e:
        st.error(f"Errore durante l'eliminazione del file {filename}: {e}")

# SOSTITUISCI LA TUA get_current_config CON QUESTA VERSIONE
def get_current_config():
    """
    Genera la configurazione dinamica. Se un codice studio è selezionato,
    crea percorsi specifici per quel cliente.
    """
    mode = st.session_state.get('tipo_struttura', 'Ditta').lower()
    codice_studio = st.session_state.get('codice_studio_valore_sicuro')

    # Percorso base per la modalità (es. ./data/ditta/)
    mode_data_dir = os.path.join(DATA_DIR, mode)
    mode_mapping_dir = os.path.join(MAPPING_BASE_DIR, mode)
    mode_export_dir = os.path.join(EXPORT_BASE_DIR, mode)

    # I percorsi diventano specifici per il cliente SE un codice studio è stato caricato
    if codice_studio:
        appoggio_dir = os.path.join(mode_data_dir, codice_studio, 'appoggio')
        mapping_dir = os.path.join(mode_mapping_dir, codice_studio)
        export_dir = os.path.join(mode_export_dir, codice_studio)
    else:
        # Se nessun cliente è selezionato, punta alle cartelle base
        appoggio_dir = os.path.join(mode_data_dir, 'appoggio')
        mapping_dir = mode_mapping_dir
        export_dir = mode_export_dir

    config = {
        "mode": mode,
        # La struttura è SEMPRE condivisa, come richiesto
        "struttura_dir": os.path.join(mode_data_dir, 'struttura'),
        # Gli altri percorsi sono dinamici
        "appoggio_dir": appoggio_dir,
        "mapping_dir": mapping_dir,
        "export_dir": export_dir,
        "db_struttura_prefix": f"struttura_{mode}_",
        "db_appoggio_suffix": f"_appoggio_{mode}"
    }
    
    # Crea tutte le cartelle necessarie per evitare errori
    for key, path in config.items():
        if key.endswith("_dir"):
            os.makedirs(path, exist_ok=True)
            
    return config

# Funzione helper per il salvataggio della configurazione globale (definita qui, non all'interno di step_5)
# Questa funzione ora gestirà una struttura di mappatura annidata
def save_global_mapping_config(config, current_full_mapping_data):
    """Salva le configurazioni globali di mappatura in file JSON."""
    mapping_dir = config["mapping_dir"]
    
    # current_full_mapping_data['column_mappings'] è ora un dizionario annidato
    global_mapping_path = os.path.join(mapping_dir, "global_mapping.json")
    with open(global_mapping_path, 'w', encoding='utf-8') as f:
        json.dump(current_full_mapping_data['column_mappings'], f, indent=4) # Salva solo il sottodizionario delle mappature
    
    date_columns_path = os.path.join(mapping_dir, "date_columns.json")
    with open(date_columns_path, 'w', encoding='utf-8') as f:
        json.dump({"date_columns": current_full_mapping_data['date_format_columns']}, f, indent=4)

    studio_mapping_path = os.path.join(mapping_dir, "studio_mapping.json")
    with open(studio_mapping_path, 'w', encoding='utf-8') as f:
        json.dump({"codice_studio_column": current_full_mapping_data['studio_code_column']}, f, indent=4)
    
    force_1to1_tables_path = os.path.join(mapping_dir, "force_1to1_tables.json")
    with open(force_1to1_tables_path, 'w', encoding='utf-8') as f:
        json.dump({"force_1to1_tables": current_full_mapping_data['force_1to1_tables']}, f, indent=4)
    
    # st.success("Configurazioni globali salvate automaticamente.") # Non mostrare in ogni rerun

# NUOVA FUNZIONE GLOBALE PER LA CALLBACK ON_CHANGE
# La callback ora gestisce una lista di valori selezionati
def update_live_mapping_callback(source_key, mode): # Rimosso current_selected_struttura_table_name dall'args
    """Aggiorna lo stato della mappatura live quando un multiselect cambia."""
    live_mapping_state_key = f"live_mapping_state_{mode}" # Revertito a chiave globale per mode
    widget_key = f"map_global_{source_key}_{mode}"
    
    if widget_key in st.session_state:
        # st.multiselect restituisce una lista, quindi salviamo la lista direttamente
        st.session_state[live_mapping_state_key][source_key] = st.session_state[widget_key]

# In pages/1_Wizard_Dati.py, nel blocco delle FUNZIONI HELPER
import unicodedata

def sanitize_column_name(col_name):
    """
    Pulisce aggressivamente il nome di una colonna:
    - Rimuove accenti e caratteri speciali.
    - Converte in minuscolo.
    - Sostituisce spazi e punteggiatura con un singolo trattino basso.
    """
    s = ''.join(c for c in unicodedata.normalize('NFD', str(col_name)) if unicodedata.category(c) != 'Mn')
    s = ''.join(c if c.isalnum() else ' ' for c in s.lower())
    return '_'.join(s.split())

# --- FUNZIONI DEGLI STEP DEL WIZARD ---

# SOSTITUISCI IL TUO step_0 CON QUESTA VERSIONE
def step_0_impostazioni(config, engine):
    mode_name = config['mode'].capitalize()
    st.header(f'Step 1: Selezione del Progetto di Lavoro ({mode_name})')
    st.info(f"Il wizard è in esecuzione in modalità: **{mode_name}**.")
    st.warning("Per cambiare modalità, torna alla pagina di Configurazione Iniziale.")
    codice_studio = st.session_state.get('codice_studio_valore_sicuro')

 # --- Funzione helper per la pulizia completa ---
    def _svuota_database_e_resetta_stato():
        """
        CANCELLA FISICAMENTE TUTTE LE TABELLE DAL DATABASE E RESETTA LO STATO DELLA SESSIONE.
        """
        try:
            with st.spinner("Cancellazione di tutte le tabelle dal database in corso..."):
                inspector = inspect(engine)
                all_tables = inspector.get_table_names()
                with engine.connect() as connection:
                    with connection.begin() as transaction:
                        # Disabilita temporaneamente i vincoli per SQLite per evitare errori di dipendenza
                        connection.execute(text("PRAGMA foreign_keys = OFF;"))
                        for table_name in all_tables:
                            connection.execute(text(f'DROP TABLE IF EXISTS "{table_name}"'))
                        # Riabilita i vincoli
                        connection.execute(text("PRAGMA foreign_keys = ON;"))
                        transaction.commit()
            
            # Pulisce lo stato della sessione, mantenendo solo le impostazioni di base
            st.success("Database svuotato con successo!")
            st.info("Reset dello stato dell'applicazione...")
            chiavi_da_mantenere = ['wizard_step', 'tipo_struttura']
            for key in list(st.session_state.keys()):
                if key not in chiavi_da_mantenere:
                    del st.session_state[key]
            
            # Resetta lo step del wizard a 0
            st.session_state.wizard_step = 0
            
        except Exception as e:
            st.error(f"Errore durante la pulizia del database: {e}")
        
        # Forza un refresh della pagina
        st.rerun()

    # --- CASO 1: UN PROGETTO È GIÀ STATO CARICATO ---
    if codice_studio:
        st.success(f"Sei attualmente al lavoro per lo studio: **{codice_studio}**")
        st.info("Tutte le operazioni avverranno nelle cartelle dedicate a questo studio. Per cambiare, chiudi prima il lavoro corrente.")
        
        def chiudi_lavoro():
            # Pulisce lo stato per permettere di cambiare studio
            keys_to_clear = [k for k in st.session_state.keys() if k != 'tipo_struttura' and k != 'wizard_step']
            for key in keys_to_clear:
                del st.session_state[key]
        
        if st.button("↩️ Chiudi Lavoro e Seleziona un altro Studio"):
            chiudi_lavoro()
            st.rerun()

    # --- CASO 2: NESSUN PROGETTO CARICATO, MOSTRIAMO LA SCHERMATA DI SCELTA ---
    else:
        st.info("Scegli se caricare un lavoro esistente o crearne uno nuovo, poi clicca 'Procedi'.")
        
        # Scansiona le cartelle per trovare i lavori già creati
        base_mapping_dir = config['mapping_dir']
        try:
            existing_studios = [d for d in os.listdir(base_mapping_dir) if os.path.isdir(os.path.join(base_mapping_dir, d))]
        except FileNotFoundError:
            existing_studios = []
        
        # Bottoni radio per scegliere l'azione
        action = st.radio(
            "Azione:",
            ["Crea un Nuovo Lavoro", "Carica un Lavoro Esistente"],
            key="project_action_radio",
            horizontal=True,
            label_visibility="collapsed"
        )

        studio_code_to_process = ""

        # Mostra l'interfaccia giusta in base alla scelta
        if action == "Carica un Lavoro Esistente":
            if not existing_studios:
                st.warning("Nessun lavoro esistente trovato. Devi prima crearne uno.")
            else:
                studio_code_to_process = st.selectbox(
                    "Scegli lo studio da caricare:",
                    options=sorted(existing_studios),
                    key="studio_loader_select"
                )
        else: # Crea un Nuovo Lavoro
            studio_code_to_process = st.text_input(
                "Inserisci il nuovo Codice Studio (3 caratteri):",
                max_chars=3,
                key="studio_creator_input"
            ).upper()

        # Unico bottone per confermare l'azione
        if st.button("Procedi", type="primary"):
            if studio_code_to_process:
                code_to_set = str(studio_code_to_process).strip()
                # Valida il codice solo se stiamo creando un nuovo lavoro
                if action == "Crea un Nuovo Lavoro" and (len(code_to_set) != 3 or not code_to_set.isalnum()):
                    st.error("Il nuovo codice deve essere di 3 caratteri alfanumerici.")
                else:
                    st.session_state.codice_studio_valore_sicuro = code_to_set
                    # Forza la creazione delle cartelle se non esistono
                    get_current_config()
                    st.success(f"Lavoro per lo studio '{code_to_set}' caricato/creato. La pagina si aggiornerà.")
                    import time
                    time.sleep(1) # Piccola pausa per far leggere il messaggio
                    st.rerun()
            else:
                st.error("Per favor, seleziona uno studio o inserisci un nuovo codice prima di procedere.")

    st.markdown("---")
    # --- Pulsante di pulizia potenziato ---
    st.warning("ATTENZIONE: L'opzione seguente cancellerà l'INTERO database (tutte le tabelle e i dati) e resetterà la sessione di lavoro.")
    if st.button('🗑️ Svuota INTERO Database e Resetta', key='svuota_db_top', on_click=_svuota_database_e_resetta_stato):
        pass # La logica è gestita dalla callback on_click


def step_1_upload_struttura(config, engine):
    mode_name = config['mode'].capitalize()
    st.header(f"Step 2: Carica File Struttura ({mode_name})")
    with st.form(key=f"upload_form_struttura_{mode_name}", clear_on_submit=True):
        st.write(f"Scegli i file struttura per {mode_name}:")
        uploaded_files = st.file_uploader("Carica file struttura", type=['xlsx'], accept_multiple_files=True, label_visibility="collapsed")
        submitted = st.form_submit_button("Carica i file selezionati")
        if submitted and uploaded_files:
            for f in uploaded_files:
                with open(os.path.join(config["struttura_dir"], f.name), 'wb') as file: file.write(f.getbuffer())
            st.success(f"{len(uploaded_files)} file caricati con successo.")
    st.markdown("---")
    st.write("File attualmente presenti:")
    try:
        files = [f for f in os.listdir(config["struttura_dir"]) if f.endswith('.xlsx')]
        if not files: st.info("Nessun file presente.")
        for i, filename in enumerate(files):
            c1, c2 = st.columns([4, 1])
            c1.info(filename)
            c2.button("🗑️ Rimuovi", key=f"remove_struttura_{mode_name}_{i}", on_click=delete_file, args=(config["struttura_dir"], filename))
    except FileNotFoundError: st.warning("Cartella non ancora creata.")

# SOSTITUISCI IL TUO step_2 CON QUESTA VERSIONE "INTELLIGENTE"
def step_2_import_struttura(config, engine):
    mode_name = config['mode'].capitalize()
    st.header(f"Step 3: Importa Struttura ({mode_name})")
    st.info(f"I file vengono letti da: `{config['struttura_dir']}`")
    try:
        files = [f for f in os.listdir(config["struttura_dir"]) if f.endswith('.xlsx')]
    except FileNotFoundError: 
        st.error("Cartella struttura non trovata."); return
    if not files: 
        st.warning('Nessun file .xlsx trovato.'); return
    
    selected_files = st.multiselect('Seleziona file da importare', files, default=files, key=f'struttura_ms_{mode_name}')
    numeric_header_row = st.number_input("Riga intestazioni NUMERICHE", min_value=1, value=2, key=f'struttura_numeric_header_{mode_name}')
    desc_header_row = st.number_input("Riga intestazioni DESCRITTIVE", min_value=1, value=3, key=f'struttura_desc_header_{mode_name}')
    
    if st.button('Importa Struttura', key=f'importa_struttura_btn_{mode_name}'):
        with st.spinner("Importazione in corso..."):
            
            # --- BLOCCO AGGIUNTO: PULIZIA PREVENTIVA DELLE VECCHIE TABELLE STRUTTURA ---
            try:
                st.write("Pulizia delle vecchie tabelle di struttura...")
                inspector = inspect(engine)
                all_db_tables = inspector.get_table_names()
                struttura_tables_to_drop = [t for t in all_db_tables if t.startswith(config["db_struttura_prefix"])]
                
                if struttura_tables_to_drop:
                    with engine.connect() as connection:
                        with connection.begin() as transaction:
                            for table_name in struttura_tables_to_drop:
                                connection.execute(text(f'DROP TABLE IF EXISTS "{table_name}"'))
                            transaction.commit()
                    st.info(f"Rimosse {len(struttura_tables_to_drop)} vecchie tabelle di struttura.")
                else:
                    st.info("Nessuna vecchia tabella di struttura da rimuovere.")
            except Exception as e:
                st.error(f"Errore durante la pulizia delle vecchie tabelle di struttura: {e}")
                st.stop()
            # --- FINE BLOCCO DI PULIZIA ---

            for file_name in selected_files:
                try:
                    workbook = openpyxl.load_workbook(os.path.join(config["struttura_dir"], file_name), read_only=True)
                    sheet = workbook.active
                    numeric_values = [cell.value for cell in sheet[numeric_header_row]]
                    descriptive_values = [cell.value for cell in sheet[desc_header_row]]
                    if descriptive_values and str(descriptive_values[0]).strip().lower() == 'non modificare questa riga':
                        numeric_values.pop(0); descriptive_values.pop(0)
                    
                    header_map = {}; final_headers = []; pretty_name_map = {}
                    for desc, num in zip(descriptive_values, numeric_values):
                        if desc and str(desc).strip():
                            original_desc = ' '.join(str(desc).strip().split())
                            clean_desc = sanitize_column_name(original_desc)
                            final_headers.append(clean_desc)
                            header_map[clean_desc] = num
                            pretty_name_map[clean_desc] = original_desc
                    
                    df_structure = pd.DataFrame(columns=final_headers)
                    table_name = f'{config["db_struttura_prefix"]}{os.path.splitext(file_name)[0]}'
                    # Nota: if_exists='replace' qui agisce come 'create' perché abbiamo già cancellato tutto
                    df_structure.to_sql(table_name, engine, if_exists='replace', index=False)
                    st.success(f"Struttura '{table_name}' importata con successo.")
                    
                    # Salva le mappe dei nomi per l'export e la UI
                    pretty_name_map_path = os.path.join(config["mapping_dir"], f"{table_name}_prettynames.json")
                    with open(pretty_name_map_path, 'w', encoding='utf-8') as f: json.dump(pretty_name_map, f, indent=4)
                    
                    # Salva la mappa per le intestazioni numeriche (potrebbe servire in futuro per l'export)
                    header_map_path = os.path.join(config["mapping_dir"], f"{table_name}_headers.json")
                    with open(header_map_path, 'w', encoding='utf-8') as f: json.dump(header_map, f, indent=4)

                except Exception as e: 
                    st.error(f"Errore importando {file_name}: {e}")

def step_3_upload_appoggio(config, engine):
    mode_name = config['mode'].capitalize()
    st.header(f'Step 4: Carica File Appoggio ({mode_name})')
    with st.form(key=f"upload_form_appoggio_{mode_name}", clear_on_submit=True):
        st.write(f"Scegli i file di appoggio per {mode_name}:")
        uploaded_files = st.file_uploader("Carica file di appoggio", type=['xlsx'], accept_multiple_files=True, label_visibility="collapsed")
        submitted = st.form_submit_button("Carica i file selezionati")
        if submitted and uploaded_files:
            for f in uploaded_files:
                with open(os.path.join(config["appoggio_dir"], f.name), 'wb') as file: file.write(f.getbuffer())
            st.success(f"{len(uploaded_files)} file caricati.")
    st.markdown("---")
    st.write("File attualmente presenti:")
    try:
        files = [f for f in os.listdir(config["appoggio_dir"]) if f.endswith('.xlsx')]
        if not files: st.info("Nessun file presente.")
        for i, filename in enumerate(files):
            c1, c2 = st.columns([4, 1])
            c1.info(filename)
            c2.button("🗑️ Rimuovi", key=f"remove_appoggio_{mode_name}_{i}", on_click=delete_file, args=(config["appoggio_dir"], filename))
    except FileNotFoundError: st.warning("Cartella non ancora creata.")

# SOSTITUISCI IL TUO step_4 CON QUESTA VERSIONE "INTELLIGENTE"
def step_4_import_appoggio(config, engine):
    mode_name = config['mode'].capitalize()
    st.header(f"Step 5: Importa Dati di Appoggio ({mode_name})")
    try:
        files = [f for f in os.listdir(config["appoggio_dir"]) if f.endswith('.xlsx')]
    except FileNotFoundError: 
        st.error("Cartella di appoggio non trovata."); return
    if not files: 
        st.warning('Nessun file trovato.'); return
    
    selected_files = st.multiselect('Seleziona file da importare', files, default=files, key=f'appoggio_ms_{mode_name}')
    header_row = st.number_input("Riga intestazioni", min_value=1, value=1, key=f'appoggio_header_{mode_name}')
    
    if st.button('Importa Dati', key=f'importa_appoggio_btn_{mode_name}'):
        with st.spinner("Importazione in corso..."):
            
            # --- BLOCCO AGGIUNTO: PULIZIA PREVENTIVA DELLE VECCHIE TABELLE DI APPOGGIO ---
            try:
                st.write("Pulizia delle vecchie tabelle di appoggio...")
                inspector = inspect(engine)
                all_db_tables = inspector.get_table_names()
                appoggio_tables_to_drop = [t for t in all_db_tables if t.endswith(config["db_appoggio_suffix"])]
                
                if appoggio_tables_to_drop:
                    with engine.connect() as connection:
                        with connection.begin() as transaction:
                            for table_name in appoggio_tables_to_drop:
                                connection.execute(text(f'DROP TABLE IF EXISTS "{table_name}"'))
                            transaction.commit()
                    st.info(f"Rimosse {len(appoggio_tables_to_drop)} vecchie tabelle di appoggio.")
                else:
                    st.info("Nessuna vecchia tabella di appoggio da rimuovere.")
            except Exception as e:
                st.error(f"Errore durante la pulizia delle vecchie tabelle di appoggio: {e}")
                st.stop()
            # --- FINE BLOCCO DI PULIZIA ---

            for file_name in selected_files:
                try:
                    file_path = os.path.join(config["appoggio_dir"], file_name)
                    
                    # Estrazione commenti con openpyxl (logica invariata)
                    workbook = openpyxl.load_workbook(file_path)
                    sheet = workbook.active
                    comments_map = {}
                    for cell in sheet[header_row]:
                        if cell.comment and cell.value:
                            sanitized_header = sanitize_column_name(cell.value)
                            raw_text = cell.comment.text
                            colon_position = raw_text.find(':')
                            comment_text = raw_text[colon_position + 1:].strip() if colon_position != -1 else raw_text.strip()
                            comments_map[sanitized_header] = comment_text
                    
                    comments_path = os.path.join(config["mapping_dir"], "appoggio_comments.json")
                    with open(comments_path, 'w', encoding='utf-8') as f:
                        json.dump(comments_map, f, indent=4)
                    if comments_map:
                        st.success(f"Trovati e salvati {len(comments_map)} commenti da `{file_name}`.")

                    # Lettura dati e importazione nel DB
                    df = pd.read_excel(file_path, header=header_row - 1, dtype=str).fillna('')
                    pretty_name_map = {sanitize_column_name(col): str(col).strip() for col in df.columns}
                    df.columns = [sanitize_column_name(col) for col in df.columns]

                    table_name = f'{os.path.splitext(file_name)[0]}{config["db_appoggio_suffix"]}'
                    df.to_sql(table_name, engine, if_exists='replace', index=False)
                    st.success(f"Dati '{table_name}' importati con colonne sanificate.")

                    pretty_name_map_path = os.path.join(config["mapping_dir"], f"{table_name}_prettynames.json")
                    with open(pretty_name_map_path, 'w', encoding='utf-8') as f: json.dump(pretty_name_map, f, indent=4)
                    
                except Exception as e: 
                    st.error(f"Errore importando {file_name}: {e}")

# SOSTITUISCI INTERAMENTE LA TUA FUNZIONE step_5_mappatura_globale CON QUESTA
def step_5_mappatura_globale(config, engine):
    mode_name = config['mode'].capitalize()
    mode = config['mode']
    st.header(f'Step 6: Mappatura Globale ({mode_name})')

    # current_full_mapping_data sarà popolato alla fine della funzione
    current_full_mapping_data = {
        "column_mappings": {}, 
        "date_format_columns": [],
        "studio_code_column": "", 
        "force_1to1_tables": []
    }

    try:
        # --- 1. CARICAMENTO DATI E SETUP ---
        inspector = inspect(engine)
        all_tables = inspector.get_table_names()
        struttura_tables = sorted([t for t in all_tables if t.startswith(config["db_struttura_prefix"])])
        appoggio_tables = sorted([t for t in all_tables if t.endswith(config["db_appoggio_suffix"])])
        
        if not (struttura_tables and appoggio_tables):
            st.error(f'Importa tabelle Struttura e Appoggio per la modalità {mode_name}.'); return
        
        # Carica pretty names per tutte le tabelle e colonne
        all_source_cols_sanitized_full_paths = [] 
        master_pretty_name_map = {} 
        for table_name_sanitized in (struttura_tables + appoggio_tables):
            path = os.path.join(config["mapping_dir"], f"{table_name_sanitized}_prettynames.json")
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    master_pretty_name_map.update(json.load(f))
            display_table_name_formatted = table_name_sanitized.replace(config['db_struttura_prefix'], '').replace(config['db_appoggio_suffix'], '').replace('_', ' ').strip()
            master_pretty_name_map[table_name_sanitized] = display_table_name_formatted

        # Raccogli tutte le colonne sorgente con il loro percorso completo
        for appoggio_tbl in appoggio_tables:
            cols_in_appoggio_tbl = pd.read_sql_table(appoggio_tbl, engine).columns.tolist()
            for col in cols_in_appoggio_tbl:
                all_source_cols_sanitized_full_paths.append(f"{appoggio_tbl}.{col}")
        source_cols_for_ui = sorted(list(set(all_source_cols_sanitized_full_paths)))

        # Funzione helper per ottenere nomi leggibili
        def get_pretty_name(s_name):
            return master_pretty_name_map.get(s_name, s_name)
        
        # --- NUOVA LOGICA PER OPZIONI DI DESTINAZIONE (MAPPATURA ASTRATTA) ---
        # Raccogliamo tutti i nomi di colonna UNICI da tutte le tabelle di struttura
        unique_dest_col_names_sanitized = set()
        for table_name in struttura_tables:
            cols_in_struttura_tbl = pd.read_sql(f'SELECT * FROM "{table_name}" LIMIT 0', engine).columns.tolist()
            for col_name_sanitized in cols_in_struttura_tbl:
                unique_dest_col_names_sanitized.add(col_name_sanitized)
        
        # Questa sarà la nostra lista di opzioni per la mappatura
        sorted_unique_dest_col_names = sorted(list(unique_dest_col_names_sanitized))
        dest_options_for_multiselect = ["-- Non mappare --", "Nascondi"] + sorted_unique_dest_col_names
        
        # NUOVA Funzione per formattare le opzioni di destinazione (mostra solo il pretty name)
        def format_dest_col_name_for_display(sanitized_col_name):
            if sanitized_col_name in ["-- Non mappare --", "Nascondi"]:
                return sanitized_col_name
            return get_pretty_name(sanitized_col_name)

        # Caricamento commenti dalle intestazioni dei file di appoggio
        comments_path = os.path.join(config["mapping_dir"], "appoggio_comments.json")
        comments_map = json.load(open(comments_path, 'r', encoding='utf-8')) if os.path.exists(comments_path) else {}

        # --- 2. GESTIONE STATO E CARICAMENTO MAPPATURE ESISTENTI ---
        loaded_full_mapping_data_from_file = {}
        mapping_path = os.path.join(config["mapping_dir"], "global_mapping.json")
        if os.path.exists(mapping_path):
            with open(mapping_path, 'r', encoding='utf-8') as f:
                loaded_full_mapping_data_from_file = json.load(f)
        
        loaded_global_settings = st.session_state.get('loaded_template_data', {})
        if not loaded_global_settings: 
            date_columns_path = os.path.join(config["mapping_dir"], "date_columns.json")
            if os.path.exists(date_columns_path):
                with open(date_columns_path, 'r', encoding='utf-8') as f:
                    loaded_global_settings.setdefault("date_format_columns", json.load(f).get("date_columns", []))
            
            studio_mapping_path = os.path.join(config["mapping_dir"], "studio_mapping.json")
            if os.path.exists(studio_mapping_path):
                with open(studio_mapping_path, 'r', encoding='utf-8') as f:
                    loaded_global_settings.setdefault("studio_code_column", json.load(f).get("codice_studio_column", ""))

            force_1to1_tables_path = os.path.join(config["mapping_dir"], "force_1to1_tables.json")
            if os.path.exists(force_1to1_tables_path):
                with open(force_1to1_tables_path, 'r', encoding='utf-8') as f:
                    loaded_global_settings.setdefault("force_1to1_tables", json.load(f).get("force_1to1_tables", []))
            else: 
                loaded_global_settings.setdefault("force_1to1_tables", [])
        
        # Inizializzazione dello stato live della mappatura
        live_mapping_state_key = f"live_mapping_state_{mode}" 
        if live_mapping_state_key not in st.session_state:
            st.session_state[live_mapping_state_key] = {}
            for col_full_path in source_cols_for_ui:
                default_dest_list = loaded_full_mapping_data_from_file.get(col_full_path, ["-- Non mappare --"])
                if not isinstance(default_dest_list, list):
                    default_dest_list = [default_dest_list]
                # Valida che le opzioni caricate esistano ancora
                valid_default_dest_list = [d for d in default_dest_list if d in dest_options_for_multiselect]
                st.session_state[live_mapping_state_key][col_full_path] = valid_default_dest_list if valid_default_dest_list else ["-- Non mappare --"]
        
        current_live_mapping = st.session_state[live_mapping_state_key]

        # --- 3. GESTIONE TEMPLATE (UI) ---
        st.subheader("Gestione Template di Mappatura")
        templates_dir = os.path.join(config["mapping_dir"], "templates")
        os.makedirs(templates_dir, exist_ok=True)
        saved_templates = ["-- Non caricare nulla --"] + sorted([os.path.splitext(f)[0].replace('_', ' ') for f in os.listdir(templates_dir) if f.endswith('.json')])
        
        loaded_name_for_display = st.session_state.get('loaded_template_name')
        index = saved_templates.index(loaded_name_for_display) if loaded_name_for_display in saved_templates else 0
        st.selectbox("Carica un Template:", options=saved_templates, index=index, key=f"template_loader_{mode}", on_change=load_template_callback, args=(config,))
        st.markdown("---")

        # --- 4. INTERFACCIA DI MAPPATURA ---
        st.subheader(f"Mappatura Colonne Dati Sorgente a Descrizioni di Destinazione")
        st.info("Mappa una sorgente a una descrizione. Se quella descrizione esiste in più tabelle di struttura, verranno popolate tutte automaticamente nei casi di mappatura semplice (1-a-1).")

        hide_unmapped_key = f'hide_unmapped_{mode}_global'
        hide_hidden_key = f'hide_hidden_{mode}_global'
        st.checkbox("Nascondi colonne non mappate", key=hide_unmapped_key, value=st.session_state.get(hide_unmapped_key, False))
        st.checkbox("Nascondi colonne impostate su 'Nascondi'", key=hide_hidden_key, value=st.session_state.get(hide_hidden_key, False))
        st.markdown("---")

        cols_to_display_filtered = source_cols_for_ui 
        if st.session_state[hide_unmapped_key]:
            cols_to_display_filtered = [c for c in cols_to_display_filtered if current_live_mapping.get(c) not in [["-- Non mappare --"], []]]
        if st.session_state[hide_hidden_key]:
            cols_to_display_filtered = [c for c in cols_to_display_filtered if current_live_mapping.get(c) != ["Nascondi"]]

        # Logica di paginazione
        page_key = f'mapping_page_{mode}_global'
        if page_key not in st.session_state: st.session_state[page_key] = 0
        items_per_page = 10
        total_pages = (len(cols_to_display_filtered) + items_per_page - 1) // items_per_page
        st.session_state[page_key] = min(st.session_state.get(page_key, 0), max(0, total_pages - 1))
        start_index, end_index = st.session_state[page_key] * items_per_page, (st.session_state[page_key] + 1) * items_per_page
        paginated_cols = cols_to_display_filtered[start_index:end_index]

        st.write(f"**Mappa Colonne Sorgente (Pagina {st.session_state[page_key] + 1} di {total_pages})**")
        
        for col_full_path in paginated_cols:
            default_selected_destinations = current_live_mapping.get(col_full_path, ["-- Non mappare --"])
            source_table_name_raw, source_col_name_sanitized = col_full_path.split('.', 1)
            source_table_display_name = get_pretty_name(source_table_name_raw)
            display_pretty_name = get_pretty_name(source_col_name_sanitized)
            comment_text = comments_map.get(source_col_name_sanitized)
            label_to_show = f"`{display_pretty_name}` (da `{source_table_display_name}`) → 💬" if comment_text else f"`{display_pretty_name}` (da `{source_table_display_name}`) →"

            st.multiselect(
                label_to_show,
                options=dest_options_for_multiselect,           # Usa la nuova lista di opzioni astratte
                default=default_selected_destinations, 
                key=f"map_global_{col_full_path}_{mode}",
                on_change=update_live_mapping_callback, 
                args=(col_full_path, mode,),
                help=comment_text,
                format_func=format_dest_col_name_for_display    # Usa la nuova funzione di formattazione
            )

        if total_pages > 1:
            st.markdown("---")
            c1,c2,c3 = st.columns([2,3,2]); 
            if st.session_state[page_key] > 0: c1.button("⬅️ Prec.", on_click=lambda: st.session_state.__setitem__(page_key, st.session_state[page_key] - 1))
            c2.write(f"<div style='text-align: center;'>Pagina {st.session_state[page_key] + 1} di {total_pages}</div>", unsafe_allow_html=True)
            if st.session_state[page_key] < total_pages - 1: c3.button("Succ. ➡️", on_click=lambda: st.session_state.__setitem__(page_key, st.session_state[page_key] + 1))
        st.markdown("---")

        # --- 5. IMPOSTAZIONI AGGIUNTIVE GLOBALI ---
        st.subheader("Impostazioni Aggiuntive Globali")
        
        # Questi widget ora usano la lista di nomi di colonna unici (astratti)
        valid_default_dates = [c for c in loaded_global_settings.get('date_format_columns', []) if c in sorted_unique_dest_col_names]
        selected_date_cols = st.multiselect("Colonne da formattare come Data (gg/mm/aaaa):", 
                                              options=sorted_unique_dest_col_names, default=valid_default_dates, 
                                              key=f'date_cols_selector_{mode}', format_func=format_dest_col_name_for_display)
        
        studio_opts = ["-- Non applicare --"] + sorted_unique_dest_col_names
        default_studio = loaded_global_settings.get('studio_code_column', studio_opts[0])
        default_studio_idx = studio_opts.index(default_studio) if default_studio in studio_opts else 0
        selected_studio_col = st.selectbox("Colonna per Codice Studio:", options=studio_opts, index=default_studio_idx,
                                           key=f'studio_col_selector_{mode}', format_func=format_dest_col_name_for_display)
        
        valid_default_force_1to1 = [t for t in loaded_global_settings.get('force_1to1_tables', []) if t in struttura_tables]
        selected_force_1to1_tables = st.multiselect(
            "Forza mappatura 1-a-1 per queste tabelle di struttura:",
            options=struttura_tables, default=valid_default_force_1to1,
            key=f'force_1to1_tables_selector_{mode}', format_func=get_pretty_name,
            help="Seleziona le tabelle di struttura che devono essere popolate con una mappatura 1-a-1, ignorando la logica di trasformazione 'molti-a-uno'."
        )
        st.markdown("---")

        # --- 6. AZIONI SUL TEMPLATE E SALVATAGGIO ---
        current_full_mapping_data['column_mappings'] = { 
            s_key: d_list for s_key, d_list in current_live_mapping.items() 
            if d_list and d_list != ["-- Non mappare --"] and d_list != ["Nascondi"]
        }
        current_full_mapping_data['date_format_columns'] = selected_date_cols
        current_full_mapping_data['studio_code_column'] = selected_studio_col if selected_studio_col != "-- Non applicare --" else ""
        current_full_mapping_data['force_1to1_tables'] = selected_force_1to1_tables

        with st.expander("Azioni su Template e Salvataggio Sessione"):
            template_name_input = st.text_input("Nome per nuovo template:", key=f"template_name_input_{mode}")
            if st.button("Salva come Nuovo Template", key=f"save_as_template_btn_{mode}"):
                if template_name_input.strip():
                    safe_filename = "".join(c for c in template_name_input if c.isalnum() or c in (' ', '_')).rstrip().replace(' ', '_')
                    template_path = os.path.join(templates_dir, f"{safe_filename}.json")
                    with open(template_path, 'w', encoding='utf-8') as f:
                        json.dump(current_full_mapping_data, f, indent=4)
                    st.success(f"Template '{template_name_input}' salvato!"); st.rerun()
                else: 
                    st.error("Inserisci un nome per il template.")

            if st.button("Aggiorna Template Caricato", key=f"update_template_btn_{mode}", disabled=(not st.session_state.get('loaded_template_name'))):
                loaded_name_for_display = st.session_state.get('loaded_template_name')
                if loaded_name_for_display:
                    safe_filename = "".join(c for c in loaded_name_for_display if c.isalnum() or c in (' ', '_')).rstrip().replace(' ', '_')
                    template_path = os.path.join(templates_dir, f"{safe_filename}.json")
                    with open(template_path, 'w', encoding='utf-8') as f:
                        json.dump(current_full_mapping_data, f, indent=4)
                    st.success(f"Template '{loaded_name_for_display}' aggiornato!")

            st.markdown("---")
            st.info("Le configurazioni sono salvate automaticamente ad ogni interazione.")
    
    except Exception as e: 
        st.error(f"Errore critico durante la mappatura: {e}"); st.exception(e)

    # Salvataggio automatico ad ogni interazione
    save_global_mapping_config(config, current_full_mapping_data)


# SOSTITUISCI IL TUO STEP 5B CON QUESTA VERSIONE AGGIORNATA
def step_5b_verifica_trasformazioni(config, engine):
    st.header("Step 6b: Configura Colonne Chiave per Trasformazioni")
    st.info("Questo step analizza la mappatura astratta. Se rileva che una descrizione è mappata da più sorgenti, ti permette di configurare la trasformazione per le tabelle che la contengono.")

    try:
        # Carica la mappatura astratta {source_full_path: [dest_col_name_1, ...]}
        mapping_path = os.path.join(config["mapping_dir"], "global_mapping.json")
        if not os.path.exists(mapping_path):
            st.warning("Esegui prima la mappatura allo Step 6."); return
        with open(mapping_path, 'r', encoding='utf-8') as f:
            global_mapping_abstract = json.load(f)
        
        force_1to1_tables_path = os.path.join(config["mapping_dir"], "force_1to1_tables.json")
        force_1to1_tables = []
        if os.path.exists(force_1to1_tables_path):
            with open(force_1to1_tables_path, 'r', encoding='utf-8') as f:
                force_1to1_tables = json.load(f).get("force_1to1_tables", [])

        # Carica i nomi leggibili per un output più chiaro
        inspector = inspect(engine)
        all_tables = inspector.get_table_names()
        master_pretty_name_map = {}
        for table_name_sanitized in all_tables:
            path = os.path.join(config["mapping_dir"], f"{table_name_sanitized}_prettynames.json")
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f: master_pretty_name_map.update(json.load(f))
            display_table_name_formatted = table_name_sanitized.replace(config['db_struttura_prefix'], '').replace(config['db_appoggio_suffix'], '').replace('_', ' ').strip()
            master_pretty_name_map[table_name_sanitized] = display_table_name_formatted
        
        def get_pretty_name(s_name):
            return master_pretty_name_map.get(s_name, s_name)

        # --- NUOVA LOGICA DI RILEVAMENTO UNPIVOT ---
        # 1. Inverti la mappa per avere dest_col -> [lista di sorgenti]
        dest_to_sources_map = {}
        for source_full, dest_cols in global_mapping_abstract.items():
            for dest_col in dest_cols:
                if dest_col not in dest_to_sources_map:
                    dest_to_sources_map[dest_col] = []
                dest_to_sources_map[dest_col].append(source_full.split('.')[-1]) # Salva solo il nome della colonna sorgente

        # 2. Analizza ogni tabella struttura per vedere se è coinvolta
        struttura_tables = sorted([t for t in all_tables if t.startswith(config["db_struttura_prefix"])])
        unpivot_tables_info = {}
        for struttura_table in struttura_tables:
            if struttura_table in force_1to1_tables:
                continue

            table_dest_cols = pd.read_sql(f'SELECT * FROM "{struttura_table}" LIMIT 0', engine).columns.tolist()
            
            # Cerca le colonne in QUESTA tabella che sono target di mappature molti-a-uno o uno-a-uno
            unpivot_triggers = {}
            one_to_one_cols = []
            is_unpivot_target = False

            for dest_col in table_dest_cols:
                sources = dest_to_sources_map.get(dest_col, [])
                if len(sources) > 1:
                    unpivot_triggers[dest_col] = sources
                    is_unpivot_target = True
                elif len(sources) == 1:
                    one_to_one_cols.append(dest_col)
            
            if is_unpivot_target:
                unpivot_tables_info[struttura_table] = {"triggers": unpivot_triggers, "key_options": one_to_one_cols}

        st.markdown("---")
        
        if not unpivot_tables_info:
            st.success("✅ Nessuna trasformazione 'molti-a-uno' rilevata. Puoi procedere."); return

        st.warning(f"⚠️ Rilevata una trasformazione 'molti-a-uno' per {len(unpivot_tables_info)} tabella/e.")
        
        # Carica/Inizializza la configurazione delle chiavi dallo stato
        unpivot_keys_path = os.path.join(config["mapping_dir"], "unpivot_keys_config.json")
        if 'unpivot_keys_config' not in st.session_state:
            st.session_state.unpivot_keys_config = json.load(open(unpivot_keys_path)) if os.path.exists(unpivot_keys_path) else {}

        # Mostra l'interfaccia di configurazione
        for table_name, info in unpivot_tables_info.items():
            with st.expander(f"**Configura le chiavi per: `{get_pretty_name(table_name)}`**"):
                st.write("Causa della trasformazione:")
                for dest_col, source_cols_list in info["triggers"].items():
                    st.markdown(f"- La colonna **`{get_pretty_name(dest_col)}`** è mappata da {len(source_cols_list)} sorgenti: `{', '.join([get_pretty_name(s) for s in source_cols_list])}`.")

                st.markdown("---")
                st.write("**Azione richiesta:** Scegli quali colonne (tra quelle mappate 1-a-1) vuoi ripetere su ogni riga creata.")
                st.caption("Default: verranno usate tutte le colonne mappate 1-a-1 come chiavi.")

                key_options = info["key_options"]
                if not key_options:
                    st.warning("Questa tabella non ha colonne con mappatura 1-a-1 da usare come chiave.")
                    st.session_state.unpivot_keys_config[table_name] = []
                    continue

                default_keys = st.session_state.unpivot_keys_config.get(table_name, [])
                valid_defaults = [k for k in default_keys if k in key_options]

                selected_keys = st.multiselect(
                    "Colonne chiave da ripetere:",
                    options=key_options, default=valid_defaults,
                    key=f"unpivot_keys_{table_name}", format_func=get_pretty_name
                )
                st.session_state.unpivot_keys_config[table_name] = selected_keys

        if st.button("Salva Configurazione Colonne Chiave"):
            with open(unpivot_keys_path, 'w', encoding='utf-8') as f:
                json.dump(st.session_state.unpivot_keys_config, f, indent=4)
            st.success("Configurazione salvata!")

    except Exception as e:
        st.error(f"Errore: {e}"); st.exception(e)

# SOSTITUISCI IL TUO STEP 6 CON QUESTA VERSIONE CON LOGICA IBRIDA
def step_6_popola_dati(config, engine):
    mode_name = config['mode'].capitalize()
    st.header(f"Step 7: Popola Dati ({mode_name})")

    if st.button("APPLICA MAPPATURA E POPOLA", key=f'popola_btn_{mode_name}'):
        with st.spinner("Popolamento in corso..."):
            try:
                # 1. Caricamento Globale delle configurazioni
                mapping_path = os.path.join(config["mapping_dir"], "global_mapping.json")
                if not os.path.exists(mapping_path):
                    st.error("'global_mapping.json' non trovato."); return
                with open(mapping_path, 'r', encoding='utf-8') as f:
                    global_mapping_abstract = json.load(f)

                unpivot_keys_path = os.path.join(config["mapping_dir"], "unpivot_keys_config.json")
                unpivot_keys_config = json.load(open(unpivot_keys_path)) if os.path.exists(unpivot_keys_path) else {}
                
                studio_mapping_path = os.path.join(config["mapping_dir"], "studio_mapping.json")
                studio_target_col = json.load(open(studio_mapping_path))['codice_studio_column'] if os.path.exists(studio_mapping_path) else ""
                codice_studio_value = st.session_state.get('codice_studio_valore_sicuro', '').upper()

                force_1to1_tables_path = os.path.join(config["mapping_dir"], "force_1to1_tables.json")
                force_1to1_tables = json.load(open(force_1to1_tables_path))['force_1to1_tables'] if os.path.exists(force_1to1_tables_path) else []

                inspector = inspect(engine)
                all_appoggio_tables_in_db = [t for t in inspector.get_table_names() if t.endswith(config["db_appoggio_suffix"])]
                appoggio_dfs = {tbl: pd.read_sql_table(tbl, engine).astype(str) for tbl in all_appoggio_tables_in_db}
                
                struttura_tables = [t for t in inspector.get_table_names() if t.startswith(config["db_struttura_prefix"])]
                if not (appoggio_dfs and struttura_tables):
                    st.warning("Nessun dato di appoggio o tabella di struttura trovati."); return

                # Inverti la mappa per avere dest_col -> [lista di sorgenti complete]
                dest_to_sources_map = {}
                for source_full, dest_cols in global_mapping_abstract.items():
                    for dest_col in dest_cols:
                        if dest_col not in dest_to_sources_map:
                            dest_to_sources_map[dest_col] = []
                        dest_to_sources_map[dest_col].append(source_full)

                # 2. Ciclo di Esecuzione per ogni tabella struttura
                for struttura_table in struttura_tables:
                    st.write(f"--- Elaborazione per `{struttura_table}` ---")
                    
                    dest_cols_for_this_table = pd.read_sql(f'SELECT * FROM "{struttura_table}" LIMIT 0', engine).columns.tolist()
                    
                    # Determina se questa tabella necessita di una trasformazione unpivot
                    is_unpivot = False
                    if struttura_table not in force_1to1_tables:
                        for dest_col in dest_cols_for_this_table:
                            if len(dest_to_sources_map.get(dest_col, [])) > 1:
                                is_unpivot = True
                                break

                    df_popolato = pd.DataFrame()

                    # --- LOGICA IBRIDA ---
                    if is_unpivot:
                        st.info(f"Logica Rilevata: Trasformazione Wide-to-Long (Unpivot) per `{struttura_table}`")
                        
                        table_specific_dest_map = {k: v for k, v in dest_to_sources_map.items() if k in dest_cols_for_this_table}
                        
                        one_to_one_map = {dest: sources[0] for dest, sources in table_specific_dest_map.items() if len(sources) == 1}
                        unpivot_map = {dest: sources for dest, sources in table_specific_dest_map.items() if len(sources) > 1}

                        all_source_tables = {s.split('.')[0] for sources_list in table_specific_dest_map.values() for s in sources_list}
                        if not all_source_tables: continue
                        
                        if len(all_source_tables) > 1:
                             st.warning(f"La trasformazione per `{struttura_table}` usa dati da più tabelle sorgente. Si assume una chiave comune implicita, il che potrebbe portare a risultati inattesi.")
                        
                        # In caso di unpivot, si assume una singola tabella di appoggio principale.
                        # La logica di join per unpivot multi-tabella non è definita.
                        source_table_name = list(all_source_tables)[0]
                        df_appoggio_current = appoggio_dfs[source_table_name]
                        
                        clean_one_to_one = {dest: src.split('.')[-1] for dest, src in one_to_one_map.items()}
                        clean_unpivot = {dest: [s.split('.')[-1] for s in src_list] for dest, src_list in unpivot_map.items()}

                        user_defined_keys = unpivot_keys_config.get(struttura_table, [])
                        key_cols_map = {k: v for k, v in clean_one_to_one.items() if k in user_defined_keys} if user_defined_keys else clean_one_to_one
                        context_cols_map = {k: v for k, v in clean_one_to_one.items() if k not in user_defined_keys} if user_defined_keys else {}
                            
                        df_popolato = crea_righe_multiple(df_appoggio_current, key_cols_map, context_cols_map, clean_unpivot)

                    else: # Mappatura Semplice
                        st.info(f"Logica Rilevata: Mappatura Semplice (1-a-1) per `{struttura_table}`")
                        
                        max_len_df = max(appoggio_dfs.values(), key=len)
                        df_popolato = pd.DataFrame(index=max_len_df.index, columns=dest_cols_for_this_table)

                        for dest_col in dest_cols_for_this_table:
                            sources = dest_to_sources_map.get(dest_col, [])
                            if len(sources) == 1:
                                source_full_path = sources[0]
                                source_table, source_col = source_full_path.split('.', 1)
                                
                                if source_table in appoggio_dfs and source_col in appoggio_dfs[source_table].columns:
                                    df_popolato[dest_col] = appoggio_dfs[source_table][source_col]
                    
                    # --- APPLICAZIONE CODICE STUDIO E SALVATAGGIO ---
                    if studio_target_col and codice_studio_value and studio_target_col in df_popolato.columns:
                        df_popolato[studio_target_col] = codice_studio_value
                    
                    if not df_popolato.empty:
                        df_popolato = df_popolato.reindex(columns=dest_cols_for_this_table).fillna('')
                        df_popolato.to_sql(struttura_table, engine, if_exists='replace', index=False)
                        st.success(f"Tabella `{struttura_table}` popolata con successo con {len(df_popolato)} righe.")
                        st.dataframe(df_popolato.head())
                    else:
                        st.warning(f"Nessun dato generato per `{struttura_table}`.")

            except Exception as e: 
                st.error(f"Errore durante il popolamento: {e}"); st.exception(e)

def step_7_modifica_massiva(config, engine):
    mode_name = config['mode'].capitalize()
    st.header(f"Step 8: Modifica Massiva ({mode_name})")
    try:
        inspector = inspect(engine)
        struttura_tables = sorted([t for t in inspector.get_table_names() if t.startswith(config["db_struttura_prefix"])])
        if not struttura_tables: st.warning("Nessuna tabella dati da modificare."); return

        session_key_table = f'tabella_in_modifica_{mode_name}'; session_key_edits = f'mass_edits_{mode_name}'
        if session_key_table not in st.session_state: st.session_state[session_key_table] = ""
        
        selected_table = st.selectbox(f"1. Seleziona tabella", [""] + struttura_tables, key=f"modifica_tabella_selector_{mode_name}")
        if selected_table and selected_table != st.session_state.get(session_key_table):
            if st.button(f"Prepara Modifiche per '{selected_table}'", key=f"prepare_edit_btn_{mode_name}"):
                st.session_state[session_key_table] = selected_table
                st.session_state[session_key_edits] = [{"col": "", "val": ""}]
                st.rerun()
        elif not selected_table and st.session_state.get(session_key_table):
            st.session_state[session_key_table] = ""; st.session_state[session_key_edits] = []; st.rerun()
        
        active_table = st.session_state.get(session_key_table)
        if active_table:
            if session_key_edits not in st.session_state: st.session_state[session_key_edits] = []
            st.markdown("---"); st.subheader(f"2. Imposta le modifiche per: `{active_table}`")
            table_cols = [""] + list(pd.read_sql(f'SELECT * FROM "{active_table}" LIMIT 0', engine).columns)

            for i in range(len(st.session_state.get(session_key_edits, []))):
                with st.container(border=True):
                    c1, c2, c3 = st.columns([4, 4, 1])
                    edit = st.session_state[session_key_edits][i]
                    default_col_idx = table_cols.index(edit["col"]) if "col" in edit and edit["col"] in table_cols else 0
                    edit["col"] = c1.selectbox("Colonna", table_cols, index=default_col_idx, key=f"edit_col_{mode_name}_{i}")
                    edit["val"] = c2.text_input("Nuovo Valore", value=edit.get("val", ""), key=f"edit_val_{mode_name}_{i}")
                    if c3.button("🗑️", key=f"remove_edit_{mode_name}_{i}", help="Rimuovi"):
                        st.session_state[session_key_edits].pop(i); st.rerun()
            
            c_btn1, c_btn2, _ = st.columns([2, 2, 8])
            if c_btn1.button("➕ Aggiungi modifica", key=f"add_edit_btn_{mode_name}"):
                st.session_state[session_key_edits].append({"col": "", "val": ""}); st.rerun()
            if c_btn2.button("✅ Applica modifiche", type="primary", key=f"apply_all_edits_btn_{mode_name}"):
                with st.spinner("Applicazione..."):
                    valid_edits = [e for e in st.session_state[session_key_edits] if e.get("col")]
                    if not valid_edits: st.warning("Nessuna modifica valida."); st.stop()
                    df_to_modify = pd.read_sql_table(active_table, engine)
                    for single_edit in valid_edits: df_to_modify[single_edit["col"]] = single_edit["val"]
                    df_to_modify.to_sql(active_table, engine, if_exists='replace', index=False)
                    st.success(f"Tabella '{active_table}' aggiornata!"); st.rerun()
            
            st.markdown("---"); st.write(f"Anteprima di **{active_table}**:")
            st.dataframe(pd.read_sql_table(active_table, engine))
    except Exception as e: st.error(f"Errore: {e}"); st.exception(e)

# SOSTITUISCI LA VECCHIA FUNZIONE CON QUESTA VERSIONE CORRETTA
def step_8_export_globale(config, engine):
    mode_name = config['mode'].capitalize()
    mode = config['mode']
    st.header(f'Step 9: Export Globale Finale ({mode})') # Utilizza 'mode' per key e nome file

    export_state_key = f'exported_file_paths_{mode}'
    if export_state_key not in st.session_state: 
        st.session_state[export_state_key] = None

    # --- MODIFICA 1: Sposta il checkbox qui in alto ---
    # In questo modo viene sempre visualizzato, permettendo all'utente di impostarlo
    # prima di avviare l'export o tra un export e l'altro.
    st.info("Imposta le opzioni desiderate prima di avviare l'export.")
    st.checkbox(
        "Rimuovi colonne vuote dall'export", 
        value=True, 
        key=f"export_remove_empty_cols_{mode}",
        help="Se selezionato, le colonne che non contengono alcun dato (oltre alle intestazioni) non verranno incluse nel file Excel finale."
    )
    st.markdown("---")

    # Ora gestiamo la visualizzazione dei download o del bottone di avvio
    if st.session_state.get(export_state_key):
        # --- BLOCCO VISUALIZZAZIONE DOWNLOAD (invariato) ---
        st.success(f"Export completato con successo. {len(st.session_state[export_state_key])} file sono pronti.")
        for f_path in st.session_state[export_state_key]:
            file_name = os.path.basename(f_path)
            if os.path.exists(f_path):
                with open(f_path, 'rb') as f:
                    st.download_button(
                        f"⬇️ Scarica {file_name}", 
                        f.read(), 
                        file_name, 
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                        key=f"dl_{file_name}_{mode}"
                    )
            else:
                st.error(f"File di export '{file_name}' non trovato. Riprova l'export.")
        
        if len(st.session_state[export_state_key]) > 1:
            st.markdown("---")
            with st.spinner("Creazione Archivio ZIP..."):
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w") as zipf:
                    for f_path in st.session_state[export_state_key]:
                        if os.path.exists(f_path): 
                            zipf.write(f_path, arcname=os.path.basename(f_path))
                st.download_button("📦 Scarica tutto (ZIP)", zip_buffer.getvalue(), f"export_{mode}.zip", "application/zip", key=f"dl_zip_btn_{mode}")
                
        st.markdown("---")
        if st.button("Esegui un nuovo export", key=f"clear_export_btn_{mode}"):
            st.session_state[export_state_key] = None
            st.rerun()
            
    else:
        # --- BLOCCO AVVIO EXPORT ---
        # Il checkbox è già stato disegnato sopra, qui mettiamo solo il bottone
        if st.button("AVVIA EXPORT FINALE", key=f'start_final_export_btn_{mode}', type="primary"):
            with st.spinner("Creazione file in corso..."):
                try:
                    # --- La logica interna rimane la stessa ---
                    date_format_path = os.path.join(config["mapping_dir"], "date_columns.json"); colonne_data = []
                    if os.path.exists(date_format_path):
                        with open(date_format_path, 'r', encoding='utf-8') as f: 
                            colonne_data = json.load(f).get("date_columns", [])
                    
                    inspector = inspect(engine)
                    struttura_tables = [t for t in inspector.get_table_names() if t.startswith(config["db_struttura_prefix"])]
                    if not struttura_tables: 
                        st.warning("Nessuna tabella dati da esportare trovata.")
                        st.stop()

                    generated_paths = []
                    # Nessuna esclusione implicita per cognome/nome. Verranno rimosse se vuote.

                    for struttura_table in struttura_tables:
                        base_name = struttura_table.replace(config["db_struttura_prefix"], '')
                        st.write(f"Elaborazione di `{base_name}`...")

                        header_map_path = os.path.join(config["mapping_dir"], f"{struttura_table}_headers.json"); header_map = {}
                        if os.path.exists(header_map_path):
                            with open(header_map_path, 'r', encoding='utf-8') as f: header_map = json.load(f)
                        else:
                            st.error(f"Mappa intestazioni per {struttura_table} non trovata."); continue

                        df_to_export = pd.read_sql_table(struttura_table, engine)
                        
                        df_final_for_export = df_to_export.copy()
                        if st.session_state.get(f"export_remove_empty_cols_{mode}", False):
                            if not df_to_export.empty:
                                cols_to_drop = [
                                    col for col in df_to_export.columns 
                                    if df_to_export[col].astype(str).str.strip().eq('').all()
                                ]
                                
                                if cols_to_drop:
                                    df_final_for_export = df_to_export.drop(columns=cols_to_drop)
                                    st.info(f"In '{base_name}', rimosse {len(cols_to_drop)} colonne completamente vuote.")
                            else:
                                st.warning(f"La tabella '{base_name}' è vuota, l'export per questo file sarà vuoto.")
                        
                        for col in colonne_data:
                            if col in df_final_for_export.columns:
                                df_final_for_export[col] = pd.to_datetime(df_final_for_export[col], errors='coerce').dt.strftime('%d/%m/%Y').fillna('')
                        
                        dest_cols = list(df_final_for_export.columns)
                        numeric_headers_row = [header_map.get(col, '') for col in dest_cols]
                        
                        wb_export = openpyxl.Workbook()
                        ws_export = wb_export.active
                        
                        ws_export.append(["Non modificare questa riga", base_name.upper()])
                        ws_export.append(["Non modificare questa riga"] + numeric_headers_row)
                        ws_export.append(["Non modificare questa riga"] + dest_cols)
                        
                        for row_data_tuple in df_final_for_export.itertuples(index=False, name=None):
                            ws_export.append([""] + list(row_data_tuple))

                        export_file_name = f"{base_name}_Export.xlsx"
                        export_file_path = os.path.join(config["export_dir"], export_file_name)
                        wb_export.save(export_file_path)
                        generated_paths.append(export_file_path)
                        st.success(f"File '{export_file_name}' salvato in: `{export_file_path}`") # Messaggio di debug più chiaro

                    st.session_state[export_state_key] = generated_paths
                    st.rerun()

                except Exception as e:
                    st.error(f"Errore durante l'export: {e}"); st.exception(e)

# --- GESTIONE WIZARD E UI ---
st.title('Importazione e Mappatura Dati')
st.sidebar.title("Navigazione")

step_labels = ['Impostazioni', 'Carica Struttura', 'Importa Struttura', 'Carica Dati', 'Importa Dati', 'Mappatura', 'Verifica Trasformazioni','Popola Dati', 'Modifica Massiva', 'Export Globale']
step_functions = [
    step_0_impostazioni, step_1_upload_struttura, step_2_import_struttura,
    step_3_upload_appoggio, step_4_import_appoggio, step_5_mappatura_globale,
    step_5b_verifica_trasformazioni, # <-- NUOVO STEP INSERITO
    step_6_popola_dati, step_7_modifica_massiva, step_8_export_globale
]

current_step_idx = st.session_state.get('wizard_step', 0)
if 'wizard_nav_radio' not in st.session_state: st.session_state.wizard_nav_radio = current_step_idx

def update_step():
    st.session_state.wizard_step = st.session_state.wizard_nav_radio
st.sidebar.radio(
    "Passaggi:", range(len(step_labels)), 
    format_func=lambda x: f"Step {x+1}: {step_labels[x]}", 
    index=current_step_idx, key="wizard_nav_radio", on_change=update_step
)

config = get_current_config()
# NUOVO BLOCCO PIÙ SICURO
try:
    # Tentiamo di eseguire lo step corrente
    step_idx = st.session_state.get('wizard_step', 0)
    step_functions[step_idx](config, engine)

except Exception as e:
    # Se si verifica un errore, lo mostriamo in modo sicuro senza causare altri errori.
    # Usiamo .get() per accedere alla chiave in modo sicuro.
    step_for_error_msg = st.session_state.get('wizard_step', -1) + 1
    st.error(f"Si è verificato un errore primario nello Step {step_for_error_msg}.")
    
    # Questa linea è la più importante: stamperà il VERO errore e il suo traceback.
    st.exception(e)

st.markdown("---")
c1, c2, _ = st.columns([2, 2, 8])
if st.session_state['wizard_step'] > 0:
    if c1.button('◀️ Indietro', key='nav_indietro', use_container_width=True): 
        st.session_state['wizard_step'] -= 1; st.rerun()
if st.session_state['wizard_step'] < len(step_functions) - 1:
    if c2.button('Avanti ▶️', key='nav_avanti', use_container_width=True): 
        st.session_state['wizard_step'] += 1; st.rerun()
